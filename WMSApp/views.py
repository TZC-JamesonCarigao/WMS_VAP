# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, FormView
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Count
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import VAPProductionData, ProductData, SourceData, ProductionDescription
from .forms import VAPProductionDataForm, ImportForm
from django.core.exceptions import FieldError

from copy import copy
from django.conf import settings
import os

from openpyxl import load_workbook


class ProductionRecordListView(ListView):
    model = VAPProductionData
    template_name = 'report_list.html'
    context_object_name = 'records'
    paginate_by = 20
    ordering = ['-Date', '-TimeStart']

    def get_queryset(self):
        queryset = super().get_queryset()
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        if start_date and end_date:
            queryset = queryset.filter(Date__range=[start_date, end_date])

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['start_date'] = self.request.GET.get('start_date', '')
        context['end_date'] = self.request.GET.get('end_date', '')
        return context


class ProductionRecordCreateView(CreateView):
    model = VAPProductionData
    form_class = VAPProductionDataForm
    template_name = 'create.html'
    success_url = reverse_lazy('report_list')

    def form_valid(self, form):
        # Manually calculate total time if not set
        instance = form.save(commit=False)
        if hasattr(instance, 'calculate_total_time') and not instance.Total and instance.TimeStart and instance.TimeEnd:
            instance.Total = instance.calculate_total_time()
        
        try:
            instance.save()
            messages.success(self.request, 'Production record created successfully!')
            return redirect(self.success_url)
        except Exception as e:
            messages.error(self.request, f'Error creating record: {str(e)}')
            return self.form_invalid(form)

class ProductionRecordUpdateView(UpdateView):
    model = VAPProductionData
    form_class = VAPProductionDataForm
    template_name = 'edit.html'
    success_url = reverse_lazy('report_list')

    def form_valid(self, form):
        instance = form.save(commit=False)
        if instance.TimeStart and instance.TimeEnd:
            instance.Total = instance.calculate_total_time()
        
        try:
            instance.save()
            messages.success(self.request, 'Production record updated successfully!')
            return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, f'Error updating record: {str(e)}')
            return self.form_invalid(form)

class ProductionRecordDeleteView(DeleteView):
    model = VAPProductionData
    template_name = 'delete.html'
    success_url = reverse_lazy('report_list')

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'Production record deleted successfully!')
        return response


# views.py - update dashboard_view function
def dashboard_view(request):
    # Calculate date range (last 30 days)
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get daily production totals
    daily_production = (
        VAPProductionData.objects
        .filter(Date__range=[start_date, end_date])
        .values('Date')
        .annotate(total_production=Sum('FormaticStrokeMin'))
        .order_by('Date')
    )
    
    # Prepare trend data
    trend_dates = []
    trend_values = []
    
    for day in daily_production:
        trend_dates.append(day['Date'].strftime('%Y-%m-%d'))
        total_prod = day['total_production'] if day['total_production'] is not None else 0
        trend_values.append(float(total_prod))
    
    # Get product distribution
    product_distribution = (
        VAPProductionData.objects
        .filter(Date__range=[start_date, end_date])
        .values('Product__Product')
        .annotate(total=Sum('FormaticStrokeMin'))
        .order_by('-total')[:6]  # Top 6 products
    )
    
    context = {
        'total_records': VAPProductionData.objects.count(),
        'recent_records': VAPProductionData.objects.order_by('-Date', '-TimeStart')[:10],
        'trend_labels': trend_dates,
        'trend_values': trend_values,
        'product_labels': [item['Product__Product'] for item in product_distribution],
        'product_values': [float(item['total']) for item in product_distribution],
        'total_production': sum(trend_values),
        'avg_daily_production': round(sum(trend_values)/len(trend_values), 2) if trend_values else 0,
    }
    
    return render(request, 'dashboard.html', context)


def export_to_excel(request):
    try:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        queryset = VAPProductionData.objects.all()
        if start_date:
            queryset = queryset.filter(Date__gte=start_date)
        if end_date:
            queryset = queryset.filter(Date__lte=end_date)

        records = queryset.values(
            'ID', 'Date', 'PlantLoc', 'Shift', 'Product__Product', 'TimeStart', 'TimeEnd',
            'Total', 'ProdMinDescrip__ProdMinDes', 'BoilerTemp', 'CookingOilTemp',
            'OutputTemp', 'InputTemp', 'Source__Source', 'FormaticStrokeMin'
        )

        df = pd.DataFrame.from_records(records)

        # Map display choices
        df['PlantLoc'] = df['PlantLoc'].map(dict(VAPProductionData.PLANT_CHOICES))
        df['Shift'] = df['Shift'].map(dict(VAPProductionData.SHIFT_CHOICES))

        def convert_time_to_decimal(value):
            if isinstance(value, str) and ('hr' in value or 'min' in value):
                hours = 0
                minutes = 0
                if 'hr' in value:
                    parts = value.split('hr')
                    hours = int(parts[0].strip())
                    if 'min' in parts[1]:
                        minutes = int(parts[1].replace('min', '').strip())
                elif 'min' in value:
                    minutes = int(value.replace('min', '').strip())
                return round(hours + (minutes / 60), 2)
            return value

        def format_time(value):
            if pd.notna(value):
                try:
                    value = str(value)
                    parts = value.split(':')
                    if len(parts) >= 2:
                        hour = parts[0].zfill(2)
                        minute = parts[1].zfill(2)
                        return f"{hour}:{minute}"
                except:
                    pass
            return value

        def process_value(value):
            new_value = convert_time_to_decimal(value)
            if new_value == value:
                new_value = format_time(value)
            try:
                if isinstance(new_value, (int, float)):
                    return format(round(float(new_value), 2), '.2f')
                if isinstance(new_value, str) and new_value.replace('.', '', 1).isdigit():
                    return format(round(float(new_value), 2), '.2f')
            except:
                pass
            return new_value

        df = df.applymap(process_value)

        template_path = os.path.join(settings.BASE_DIR, 'templates', 'template.xlsx')
        wb = load_workbook(template_path)
        ws = wb.active
        start_row = 2
        start_col = 1
        template_row = ws[start_row]

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                template_cell = template_row[c_idx - start_col]
                if template_cell.has_style:
                    cell.font = copy(template_cell.font)
                    cell.border = copy(template_cell.border)
                    cell.fill = copy(template_cell.fill)
                    cell.number_format = copy(template_cell.number_format)
                    cell.protection = copy(template_cell.protection)
                    cell.alignment = copy(template_cell.alignment)

        if start_date or end_date:
            file_name = f"{start_date or 'start'}_to_{end_date or 'end'}.xlsx"
        else:
            file_name = "alldata.xlsx"

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Export failed: {str(e)}", status=500)


class ImportDataView(FormView):
    template_name = 'import.html'
    form_class = ImportForm
    success_url = reverse_lazy('report_list')

    def form_valid(self, form):
        excel_file = form.cleaned_data['excel_file']

        try:
            df = pd.read_excel(excel_file)

            for _, row in df.iterrows():
                # Get or create related objects first
                product, _ = ProductData.objects.get_or_create(Product=row['Product'])
                source, _ = SourceData.objects.get_or_create(Source=row['Source'])
                description, _ = ProductionDescription.objects.get_or_create(ProdMinDes=row['Description'])

                VAPProductionData.objects.create(
                    Date=row['Date'],
                    PlantLoc=row['Plant Location'],
                    Shift=row['Shift'],
                    Product=product,
                    TimeStart=row['Start Time'],
                    TimeEnd=row['End Time'],
                    BoilerTemp=row['Boiler Temp (°C)'],
                    CookingOilTemp=row['Oil Temp (°C)'],
                    OutputTemp=row['Output Temp (°C)'],
                    InputTemp=row['Input Temp (°C)'],
                    Source=source,
                    FormaticStrokeMin=row['Formatic Strokes/Min'],
                    ProdMinDescrip=description
                )

            messages.success(self.request, 'Data imported successfully!')
        except Exception as e:
            messages.error(self.request, f'Error importing data: {str(e)}')

        return super().form_valid(form)